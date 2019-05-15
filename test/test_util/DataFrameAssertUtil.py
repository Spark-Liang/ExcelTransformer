from decimal import Decimal
from unittest import TestCase

import openpyxl
import pandas as pd
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame


def assert_data_frame_with_same_header_and_data(
        test_case, expected_data_frame, result
):
    """

    :param test_case:
    :param DataFrame expected_data_frame:
    :param DataFrame result:
    :return:
    """
    test_case.assertEqual(
        expected_data_frame.shape, result.shape
    )
    for expected_header, result_header in zip(expected_data_frame.columns.values, result.columns.values):
        test_case.assertEqual(expected_header, result_header)
    max_x, max_y = expected_data_frame.shape
    for x in range(0, max_x):
        for y in range(0, max_y):
            test_case.assertEqual(
                expected_data_frame.iloc[x].iloc[y],
                result.iloc[x].iloc[y],
                "value in ({},{}) is not equal".format(x, y)
            )


def assert_data_in_excel_is_equal(
        test_case, expected_file_path, result_file_path, sheet_name=None, assert_func=None
):
    """

    :param TestCase test_case:
    :param str expected_file_path:
    :param str result_file_path:
    :param str sheet_name:
    :param (TestCase,Cell,Cell)->None assert_func:
    :return:
    """
    expected_wb = openpyxl.load_workbook(expected_file_path)
    result_wb = openpyxl.load_workbook(result_file_path)

    # with single sheet in the excel
    if sheet_name is not None:
        test_case.assertTrue(sheet_name in result_wb)
        assert_data_in_sheet_is_equal(
            test_case, expected_wb[sheet_name], result_wb[sheet_name], assert_func
        )
        return

    for sheet_name in expected_wb.sheetnames:
        test_case.assertTrue(sheet_name in result_wb)
        expected_sheet = expected_wb[sheet_name]
        result_sheet = result_wb[sheet_name]
        assert_data_in_sheet_is_equal(test_case, expected_sheet, result_sheet, assert_func)


def assert_data_in_sheet_is_equal(test_case, expected_sheet, result_sheet, assert_func=None):
    """
    Assert all the data in the excel sheet is equal.
    :param test_case:
    :param Worksheet expected_sheet:
    :param Worksheet result_sheet:
    :param (TestCase,Cell,Cell)->None assert_func:
    :return:
    """

    if assert_func is None:
        assert_func = assert_value_in_cell_is_equal

    test_case.assertEqual(expected_sheet.max_row, result_sheet.max_row)
    test_case.assertEqual(expected_sheet.max_column, result_sheet.max_column)

    for x in range(1, expected_sheet.max_row):
        for y in range(1, expected_sheet.max_column):
            assert_func(
                test_case,
                expected_sheet.cell(x, y),
                result_sheet.cell(x, y)
            )


def assert_value_in_cell_is_equal(test_case, expected_cell, actual_cell):
    """

    :param TestCase test_case:
    :param Cell expected_cell:
    :param Cell actual_cell:
    :return:
    """
    expected_val = expected_cell.value
    actual_val = actual_cell.value
    if isinstance(expected_val, str):
        if expected_val.isnumeric() or expected_val.isnumeric():
            test_case.assertEqual(
                Decimal(expected_val), Decimal(actual_val)
            )
    elif isinstance(expected_val, (int, Decimal)):
        test_case.assertEqual(
            Decimal(expected_val), Decimal(actual_val)
        )
    elif isinstance(expected_val, float):
        test_case.assertAlmostEqual(float(expected_val), float(actual_val))
    else:
        err_msg = "value in cell ({},{}) is not equal".format(expected_cell.row, expected_cell.column)
        test_case.assertEqual(
            expected_val, actual_val,
            err_msg
        )


def assert_data_in_data_frame_is_equal(test_case, expected_df, result_df, assert_func=None):
    """
    Assert all the data in the given data frame is equal.
    :param test_case:
    :param DataFrame expected_df:
    :param DataFrame result_df:
    :param (TestCase,Any,Any)->None assert_func:
    :return:
    """
    if assert_func is None:
        assert_func = default_assert_func
    test_case.assertEqual(
        expected_df.shape,
        result_df.shape
    )

    max_x, max_y = expected_df.shape
    for header in expected_df.columns.to_list():
        expected_col = expected_df[header]
        result_col = result_df[header]
        for row_idx in range(0, max_y):
            expected_cell = expected_col[row_idx]
            actual_cell = result_col[row_idx]
            assert_func(test_case, expected_cell, actual_cell)


def default_assert_func(test_case_in_func, expected_val, actual_val):
    if pd.isna(expected_val):
        return
    test_case_in_func.assertEqual(
        expected_val,
        actual_val
    )
