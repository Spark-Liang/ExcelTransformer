import os
import unittest
from os import path

import openpyxl
import pandas as pd

from test_util.DataFrameAssertUtil import assert_data_in_excel_is_equal
from xl_transform.writer import FileWriter as SUT


class TestWrite(unittest.TestCase):

    def setUp(self):
        self.__test_target_path = "init"

    def tearDown(self):
        if path.exists(self.__test_target_path):
            os.remove(self.__test_target_path)
            self.__test_target_path = None

    def test_write_with_horizontal_and_vertical_mapping_and_with_projection(self):
        # given
        test_name = "WriteHorizontalAndVerticalMappingAndWithProjection"
        self.__test_target_path = get_test_target_file_path(test_name)
        template_path = get_test_template_path(test_name)
        data_frame_dict = get_test_data(test_name)

        # when
        SUT.write(
            self.__test_target_path,
            template_path,
            data_frame_dict,
            None
        )

        # then
        self.assert_expected_equal_to_result(test_name)

    def test_write_data_to_two_sheet(self):
        # given
        test_name = "WriteDataIntoTwoSheet"
        self.__test_target_path = get_test_target_file_path(test_name)
        template_path = get_test_template_path(test_name)
        data_frame_dict = get_test_data(test_name)
        # when
        SUT.write(
            self.__test_target_path,
            template_path,
            data_frame_dict,
            None
        )

        # then
        self.assert_expected_equal_to_result(test_name)

    def test_raise_exception_when_area_to_write_is_contacted(self):
        # given
        test_name = "WriteFailedWhenTwoMappingAreaHasContacted"
        config_path = get_writer_config_path(test_name)
        self.__test_target_path = get_test_target_file_path(test_name)
        template_path = get_test_template_path(test_name)
        data_frame_dict = get_test_data(test_name)

        # when
        def func():
            SUT.write(
                self.__test_target_path,
                template_path,
                data_frame_dict,
                config_path
            )

        self.assertRaises(Exception, func)

    def assert_expected_equal_to_result(self, test_name):
        self.assertTrue(
            path.exists(self.__test_target_path)
        )
        assert_data_in_excel_is_equal(
            self,
            get_expected_target_file_path(test_name),
            self.__test_target_path
        )


def get_test_data_prefix(test_name):
    return path.join(
        path.dirname(__file__), "test_data", "FileWriter", test_name
    )


def get_test_data(test_name):
    file_path = path.join(
        get_test_data_prefix(test_name), "Data.xlsx"
    )
    wb = openpyxl.load_workbook(file_path, read_only=True)
    result = {}
    for sheet in wb:
        mapping_name = sheet.title
        # read value by column
        value_dict = {}
        for y in range(1, sheet.max_column + 1):
            header = sheet.cell(1, y).value
            value_list = [sheet.cell(x, y).value for x in range(2, sheet.max_row + 1)]
            value_dict[header] = value_list
        result[mapping_name] = pd.DataFrame(value_dict)
    return result


def get_test_template_path(test_name):
    return path.join(
        get_test_data_prefix(test_name),
        "Template.xlsx"
    )


def get_test_target_file_path(test_name):
    test_target_file_path = path.join(
        get_test_data_prefix(test_name),
        "TmpOut.xlsx"
    )
    if path.exists(test_target_file_path):
        os.remove(test_target_file_path)
    return test_target_file_path


def get_expected_target_file_path(test_name):
    return path.join(
        get_test_data_prefix(test_name),
        "Result.xlsx"
    )


def get_writer_config_path(test_name):
    return path.join(
        get_test_data_prefix(test_name),
        "config.json"
    )
