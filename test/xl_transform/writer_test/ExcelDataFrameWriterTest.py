import json
import os
import unittest
from datetime import datetime
from decimal import Decimal
from os import path

import openpyxl
import pandas as pd
from openpyxl.cell.cell import Cell

from test_util.DataFrameAssertUtil import assert_data_in_excel_is_equal, assert_value_in_cell_is_equal
from xl_transform.common import TemplateInfoItem, row, column
from xl_transform.common.model import Cell
from xl_transform.writer import ExcelDataFrameWriter as SUT


class TestWrite(unittest.TestCase):

    def setUp(self):
        self.__test_target_path = "init"

    def tearDown(self):
        if path.exists(self.__test_target_path):
            os.remove(self.__test_target_path)
            self.__test_target_path = None

    def test_write_with_format(self):
        # given
        test_name = "WriteWithFormat"
        config = read_writer_config(test_name)
        self.__test_target_path = get_test_target_file_path(test_name)
        target_sheet_name = "Target"
        info_item = TemplateInfoItem(
            target_sheet_name, Cell(1, 1), "TBL_1",
            ["str", "int", "date_1", "date_2", "float", "decimal"], row
        )
        date_format = "%Y/%m/%d %H:%M:%S"
        data_frame = pd.DataFrame({
            "str": [10, 0.029576325, 0.333226408],
            "int": [20, -10, 0],
            "date_1": [datetime.strptime(str_val, date_format) for str_val in
                       ["2019/01/30 00:00:00", "2019/02/28 16:58:24", "2019/12/30 08:20:47"]],
            "date_2": [datetime.strptime(str_val, date_format) for str_val in
                       ["2019/01/30 00:00:00", "2019/02/28 00:00:00", "2019/12/30 00:00:00"]],
            "float": [6.144880061, -1.195569306, 2.847225604],
            "decimal": [Decimal.from_float(x) for x in [6.144880061, -1.195569306, 2.847225604]]
        }, dtype=object)
        sut = SUT(info_item, config)

        # when
        sut.write(data_frame, self.__test_target_path)

        # then

        def assert_value_and_format_is_equal(
                test_case, expected_cell, actual_cell
        ):
            """

            :param unittest.TestCase test_case:
            :param Cell expected_cell:
            :param Cell actual_cell:
            :return:
            """
            assert_value_in_cell_is_equal(test_case, expected_cell, actual_cell)
            test_case.assertEqual(
                expected_cell.number_format,
                actual_cell.number_format,
                "format in ({},{}) is not equal".format(expected_cell.row, expected_cell.column)
            )

        self.assert_expected_equal_to_result(target_sheet_name, test_name, assert_value_and_format_is_equal)

    def test_write_with_single_mapping_and_with_projection(self):
        # given
        test_name = "WriteSingleMappingWithProjection"
        data_frame = get_test_data(test_name)
        self.__test_target_path = get_test_target_file_path(test_name)
        target_sheet_name = "Target"
        info_item = TemplateInfoItem(
            target_sheet_name, Cell(3, 3), "TBL_1",
            ["C", "B", "A"], row
        )
        sut = SUT(info_item)

        # when
        sut.write(data_frame, self.__test_target_path)

        # then
        self.assert_expected_equal_to_result(target_sheet_name, test_name)

    def test_write_vertical_header(self):
        # given
        test_name = "WriteWithVerticalHeader"
        data_frame = get_test_data(test_name)
        self.__test_target_path = get_test_target_file_path(test_name)
        target_sheet_name = "Target"
        info_item = TemplateInfoItem(
            target_sheet_name, Cell(3, 3), "TBL_1",
            ["C", "B", "A"], column
        )
        sut = SUT(info_item, None)

        # when
        sut.write(data_frame, self.__test_target_path)

        # then
        self.assert_expected_equal_to_result(target_sheet_name, test_name)

    def test_write_without_header(self):
        # given
        test_name = "WriteWithoutHeader"
        data_frame = get_test_data(test_name)
        config = read_writer_config(test_name)
        self.__test_target_path = get_test_target_file_path(test_name)
        target_sheet_name = "Target"
        info_item = TemplateInfoItem(
            target_sheet_name, Cell(3, 3), "TBL_1",
            ["C", "B", "A"], row
        )
        sut = SUT(info_item, config)

        # when
        sut.write(data_frame, self.__test_target_path)

        # then
        self.assert_expected_equal_to_result(target_sheet_name, test_name)

    def test_write_limited_rows(self):
        # given
        test_name = "WriteWithLimitedRow"
        data_frame = get_test_data(test_name)
        config = read_writer_config(test_name)
        self.__test_target_path = get_test_target_file_path(test_name)
        target_sheet_name = "Target"
        info_item = TemplateInfoItem(
            target_sheet_name, Cell(3, 3), "TBL_1",
            ["C", "B", "A"], row
        )
        sut = SUT(info_item, config)

        # when
        sut.write(data_frame, self.__test_target_path)

        # then
        self.assert_expected_equal_to_result(target_sheet_name, test_name)

    def test_will_raise_exception_when_data_not_have_the_needed_column(self):
        # given
        test_name = "WillRaiseExceptionWhenDataFrameNotProvideNeededColumn"
        data_frame = get_test_data(test_name)
        config = read_writer_config(test_name)
        self.__test_target_path = get_test_target_file_path(test_name)
        target_sheet_name = "Target"
        info_item = TemplateInfoItem(
            target_sheet_name, Cell(3, 3), "TBL_1",
            ["C", "B", "A"], row
        )
        sut = SUT(info_item, config)

        # when
        def func():
            sut.write(data_frame, self.__test_target_path)

        self.assertRaises(Exception, func)

    def assert_expected_equal_to_result(self, target_sheet_name, test_name, assert_func=None):
        self.assertTrue(
            path.exists(self.__test_target_path)
        )
        assert_data_in_excel_is_equal(
            self,
            get_expected_target_file_path(test_name),
            self.__test_target_path,
            target_sheet_name,
            assert_func=assert_func
        )


def get_test_data_prefix(test_name):
    return path.join(
        path.dirname(__file__), "test_data", "ExcelDataFrameWriter", test_name
    )


def get_test_data(test_name):
    file_path = path.join(
        get_test_data_prefix(test_name), "Data.xlsx"
    )
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet = wb.active
    # read value by column
    value_dict = {}
    for y in range(1, sheet.max_column + 1):
        header = sheet.cell(1, y).value
        value_list = [sheet.cell(x, y).value for x in range(2, sheet.max_row + 1)]
        value_dict[header] = value_list
    return pd.DataFrame(value_dict)


def get_test_target_file_path(test_name):
    return path.join(
        get_test_data_prefix(test_name),
        "TmpOut.xlsx"
    )


def get_expected_target_file_path(test_name):
    return path.join(
        get_test_data_prefix(test_name),
        "Result.xlsx"
    )


def read_writer_config(test_name):
    config_file_path = path.join(
        get_test_data_prefix(test_name),
        "config.json"
    )
    with open(config_file_path, "r") as json_file:
        return json.load(json_file)
