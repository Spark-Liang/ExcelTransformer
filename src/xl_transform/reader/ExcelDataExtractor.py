import openpyxl
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from pandas import DataFrame

from xl_transform.common import row


def extract_data_frame(
        source,
        sheet_name,
        header_list=None,
        start_row_idx=1,
        start_col_idx=1,
        data_row_direction=row,
        import_header=True,
        data_row_skip=None,
        data_rows_limit=None,
        data_column_limit=None,
        dtype=None,
        converter=None,
):
    """

    :param str or Workbook source:
    :param sheet_name:
    :param header_list:
    :param int start_row_idx: start index is 1
    :param int start_col_idx: start index is 1
    :param str data_row_direction:
        <p>
            Means the actual direction in the excel of each data row in the DataFrame. <br>
            "row" means the data in the single row of the DataFrame will locate from left to right in the excel.<br>
            "column" means the data in the single row of the DataFrame will locate from top to bottom in the excel.<br>
        </p>
    :param bool import_header:
        This parameter is used to control whether import header from data or not.<br>
        This parameter will ignored when "header_list" is not None. <br>
    :param int data_row_skip:
    :param int data_rows_limit:
    :param int data_column_limit:
        Restrict the number of column to extract.   <br>
        This parameter need to provide when "import_header" is True.
    :param dict[str,(Cell)->Any] converter:
    :param dict[str,type] dtype:
    :return:
    :rtype: DataFrame
    """
    if isinstance(source, str):
        wb = openpyxl.load_workbook(source)
        if sheet_name not in wb:
            err_msg = "The given sheet not in the Excel"
            raise Exception(err_msg)
    else:
        wb = source
    sheet = wb[sheet_name]

    # proceed header
    if header_list is None and not import_header:
        # create auto-gen header
        if data_column_limit is None:
            err_msg = "The 'data_column_limit' must provided when read_data data with auto-generated header"
            raise Exception(err_msg)
        header_list = range(0, data_column_limit)
    elif header_list is not None:
        # use the given header
        import_header = False
    else:
        # import the header from data
        if data_column_limit is None:
            err_msg = "The 'data_column_limit' must provided when read_data data with header import"
            raise Exception(err_msg)
        header_list = extract_header(sheet, start_row_idx, start_col_idx, data_row_direction, data_column_limit)

    # prepare data proceed function
    def data_proceed(header_str, cell):
        """

        :param header_str:
        :param Cell cell:
        :return:
        """
        if converter is not None:
            return convert_by_converter(header_str, cell, converter)
        if dtype is not None:
            return convert_by_type_hint(header_str, cell, dtype)
        return cell.value

    # extract data from excel
    data_dict = {}
    for d_idx in range(0, len(header_list)):
        header = header_list[d_idx]
        if data_row_direction == row:
            column_idx = start_col_idx + d_idx
            data_column_value_list = \
                [data_proceed(header, sheet.cell(row_idx, column_idx))
                 for row_idx in range(start_row_idx + 1 if import_header else start_row_idx
                                      , sheet.max_row + 1)]
        else:
            row_idx = start_row_idx + d_idx
            data_column_value_list = \
                [data_proceed(header, sheet.cell(row_idx, column_idx))
                 for column_idx in range(start_col_idx + 1 if import_header else start_col_idx
                                         , sheet.max_column + 1)]

        data_dict[header] = data_column_value_list

    # get bounded data
    def get_last_not_null_value_idx(value_list):
        for ind in range(len(value_list) - 1, 0, -1):
            if value_list[ind] is not None:
                return ind
        return 0

    last_not_all_null_row_idx = max(*[get_last_not_null_value_idx(l) for l in data_dict.values()])
    if data_rows_limit is not None and data_rows_limit < last_not_all_null_row_idx:
        row_idx_to_trim = data_rows_limit
    else:
        row_idx_to_trim = last_not_all_null_row_idx + 1
    if data_row_skip is None:
        data_row_skip = 0
    data_dict = {
        header: value_list[data_row_skip:row_idx_to_trim]
        for header, value_list in data_dict.items()
    }

    return DataFrame(data_dict, columns=header_list)


def extract_header(sheet, start_row_idx, start_col_idx, header_row_direction, data_column_limit):
    if header_row_direction == row:
        header_list = [sheet.cell(start_row_idx, start_col_idx + d_idx).value
                       for d_idx in range(0, data_column_limit)]
    else:
        header_list = [sheet.cell(start_row_idx + d_idx, start_col_idx).value
                       for d_idx in range(0, data_column_limit)]
    return header_list


def convert_by_type_hint(header, cell, dtype):
    """

    :param str header:
    :param Cell cell:
    :param dict[str,type] dtype:
    :return:
    """
    if header in dtype:
        return dtype[header](cell.value)
    return cell.value


def convert_by_converter(header, cell, converter):
    """

    :param str header:
    :param Cell cell:
    :param dict[str,(Cell)->Any] converter:
    :return:
    """
    if header in converter:
        return converter[header](cell)
    return cell.value
