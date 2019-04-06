import re

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from xl_transform.common import CellsSetUtils
from xl_transform.common.model import Cell
from xl_transform.common.model import OrthorhombicLine

global row, column

# noinspection PyRedeclaration
row, column = "row", "column"


class Template(object):

    def __init__(self, template_file_path):
        """

        :param str template_file_path:
        """
        if not template_file_path.endswith(".xlsx"):
            err_msg = "Unsupported type of file: " + template_file_path
            raise Exception(err_msg)
        self.__info_items = parse_excel_template(template_file_path)

    @property
    def info_items(self):
        return list(self.__info_items)


class TemplateInfoItem(object):
    """
    Each one of the "TemplateInfoItem" represent a single template area. This class contain the below properties to describe the template information.
    Properties
    __________
    sheet_name : str
        The sheet where the template area is located in. \n
        This value will be available only when the template in an Excel file.

    top_left_point : Cell
        This property is used to location the template area.

    mapping_name : str
        The name of the associated mapping.

    headers : list[str]
        The list of the column headers exists in the template area.

    header_direction : str
        The direction of the header. It can be "row" or "column".


    """

    def __init__(
            self,
            sheet_name, top_left_point, mapping_name, headers, header_direction=row
    ):
        """

        :param str sheet_name:
        :param Cell top_left_point:
        :param str mapping_name:
        :param list[str] headers:
        :param None or str header_direction: "row" or "column".
        """
        self.__sheet_name = sheet_name
        self.__top_left_point = top_left_point
        self.__mapping_name = mapping_name
        self.__headers = tuple(headers)
        self.__header_direction = header_direction

    @property
    def top_left_point(self):
        return self.__top_left_point

    @property
    def sheet_name(self):
        return self.__sheet_name

    @property
    def mapping_name(self):
        return self.__mapping_name

    @property
    def headers(self):
        return self.__headers

    @property
    def header_direction(self):
        return self.__header_direction

    def __eq__(self, other):
        return (self.top_left_point == other.top_left_point
                and self.sheet_name == other.sheet_name
                and self.mapping_name == other.mapping_name
                and self.headers == other.headers
                and self.header_direction == other.header_direction)

    def __hash__(self):
        return hash((
            self.top_left_point,
            self.sheet_name,
            self.mapping_name,
            self.headers,
            self.header_direction
        ))


def parse_excel_template(filename):
    """

    :param str filename:
    :return:
    :rtype: list[TemplateInfoItem]
    """
    result = []

    wb = openpyxl.load_workbook(filename, read_only=True)
    for sheet in wb:
        result.extend(
            __parse_work_sheet(sheet)
        )
    return result


def __parse_work_sheet(sheet):
    """

    :param Worksheet sheet:
    :return:
    """
    max_x, max_y = sheet.max_row, sheet.max_column
    target_cell_info_dict = {}
    for x in range(1, max_x + 1):
        for y in range(1, max_y + 1):
            cell_value = sheet.cell(x, y).value
            if isinstance(cell_value, str):
                __parse_cell_info_and_set_target_into_dict(
                    x, y, cell_value, target_cell_info_dict
                )
    target_cells = target_cell_info_dict.keys()
    list_of_contact_cells_set = \
        CellsSetUtils.separate_cells_set_into_contacted_cells_set(target_cells)
    # check if all the shape of the contacted cells are orthorhombic line.
    return __check_line_shape_and_construct_parse_result_items(
        sheet.title, list_of_contact_cells_set, target_cell_info_dict
    )


def __parse_cell_info_and_set_target_into_dict(x, y, value, target_cell_info_dict):
    pattern = r"^\s*\$\{([^}]+?)\:([^}]+?)\}\s*$"
    matchObj = re.match(pattern, value)
    if matchObj is None:
        return
    mapping_name = matchObj.group(1)
    header_name = matchObj.group(2)
    target_cell_info_dict[Cell(x, y)] = {
        "mapping_name": mapping_name,
        "header_name": header_name
    }


def __check_line_shape_and_construct_parse_result_items(
        sheet_name, contact_cells_set_list, target_cell_info_dict
):
    """

    :param list[set[Cell]] contact_cells_set_list:
    :param dict[Cell,dict[str,str]] target_cell_info_dict:
    :return:
    :rtype: list[TemplateInfoItem]
    """
    result = []
    for contact_cells_set in contact_cells_set_list:
        mapping_cells_dict = {}
        for cell in contact_cells_set:
            mapping_name = target_cell_info_dict[cell]["mapping_name"]
            mapping_cell_set = mapping_cells_dict.setdefault(
                mapping_name, set()
            )
            mapping_cell_set.add(cell)
        for mapping, cells_set_of_mapping in mapping_cells_dict.items():
            line = OrthorhombicLine(cells_set_of_mapping)
            result.append(TemplateInfoItem(
                sheet_name=sheet_name,
                top_left_point=line.cells_list[0],
                mapping_name=mapping,
                headers=[target_cell_info_dict[cell]["header_name"] for cell in line.cells_list],
                header_direction=(column if line.is_vertical else row) if line.is_vertical is not None else None
            ))
    return result
