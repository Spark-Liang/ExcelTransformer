from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from pandas import DataFrame

from xl_transform.common import row


def write_data_frame_to_workbook(
        target,
        sheet_name,
        df,
        start_row_idx=1,
        start_col_idx=1,
        data_row_direction=row,
        with_header=True,
        data_row_skip=None,
        data_rows_limit=None,
        dtype=None,
        converter=None,
):
    """

    :param Workbook target:
    :param sheet_name:
    :param DataFrame df:
    :param bool with_header:
    :param int start_row_idx: start index is 1
    :param int start_col_idx: start index is 1
    :param str data_row_direction:
    :param int data_rows_limit:
    :param int data_row_skip:
    :param converter:
    :param dtype:
    :return:
    :rtype: DataFrame
    """
    wb = target
    if sheet_name not in wb:
        wb.create_sheet(sheet_name)
    sheet = wb[sheet_name]

    header_list = df.columns.values

    # write_data header if need
    if with_header:
        if data_row_direction == row:
            header_row_idx = start_row_idx
            for d_idx in range(0, len(header_list)):
                sheet.cell(header_row_idx, start_col_idx + d_idx).value = header_list[d_idx]
        else:
            header_column_idx = start_col_idx
            for d_idx in range(0, len(header_list)):
                sheet.cell(start_row_idx + d_idx, header_column_idx).value = header_list[d_idx]

    data_row_skip = data_row_skip if data_row_skip is not None else 0
    if data_rows_limit is not None:
        data_rows_limit = min(data_rows_limit, df.shape[0])
    else:
        data_rows_limit = df.shape[0]

    # write_data data
    def update_data_func(header_str, cell, val):
        """

        :param Any val:
        :param header_str:
        :param Cell cell:
        :return:
        """
        if converter is not None:
            cell.value = convert_by_converter(header_str, val, converter)
            return
        if dtype is not None:
            cell.value = convert_by_type_hint(header_str, val, dtype)
            return
        cell.value = val

    if with_header:
        if data_row_direction == row:
            base_row_idx = start_row_idx + 1
            base_col_idx = start_col_idx
        else:
            base_row_idx = start_row_idx
            base_col_idx = start_col_idx + 1
    else:
        base_row_idx = start_row_idx
        base_col_idx = start_col_idx
    for d_idx in range(0, len(header_list)):
        header = header_list[d_idx]
        value_list = df[header]
        if data_row_direction == row:
            column_idx = base_col_idx + d_idx
            for ind in range(data_row_skip, data_rows_limit):
                row_idx = base_row_idx + ind
                update_data_func(header, sheet.cell(row_idx, column_idx), value_list[ind])
        else:
            row_idx = base_row_idx + d_idx
            for ind in range(data_row_skip, data_rows_limit):
                column_idx = base_col_idx + ind
                update_data_func(header, sheet.cell(row_idx, column_idx), value_list[ind])


def convert_by_type_hint(header, value, dtype):
    """

        :param str header:
        :param Any value:
        :param dict[str,type] dtype:
        :return:
        """
    if header in dtype:
        return dtype[header](value)
    return value


def convert_by_converter(header, value, converter):
    """

        :param str header:
        :param Any value:
        :param dict[str,(Cell)->Any] converter:
        :return:
        """
    if header in converter:
        converter_func = converter[header]
        if converter_func is not None:
            return converter_func(value)
    return value
